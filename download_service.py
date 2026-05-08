"""
download_service.py
───────────────────
Builds the combined export ZIP containing:
  - metadata.xlsx  (always)
  - attachments/   (if requested)
  - email_content/ (if requested, PDF only)

ZIP structure:
  gmail_export_<YYYYMMDD>/
    metadata.xlsx
    attachments/
      <YY MM (Mon YY)>/
        Image|PDF|Excel|Document|PPT|Other/
          <YYYYMMDD>_<original_filename>
    email_content/
      <YY MM (Mon YY)>/
          <YYYYMMDD>_<subject>.pdf

metadata.xlsx gets relative hyperlinks into the ZIP for body files and
attachment files when those sections are included.

Nothing here imports streamlit.
"""

import io
import re
import base64
import zipfile
from datetime import datetime
from email.utils import parsedate_to_datetime

from googleapiclient.discovery import build

# ── Optional deps — fail with clear messages ──────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib import colors
    _HAS_REPORTLAB = True
except ImportError:
    _HAS_REPORTLAB = False

try:
    from bs4 import BeautifulSoup
    _HAS_BS4 = True
except ImportError:
    _HAS_BS4 = False


# ─────────────────────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _safe_filename(name: str, max_len: int = 80) -> str:
    name = re.sub(r'[\\/:*?"<>|]', '_', name).strip()
    return name[:max_len] if len(name) > max_len else name


def _parse_date(raw: str) -> datetime | None:
    try:
        return parsedate_to_datetime(raw)
    except Exception:
        return None


def _month_folder(dt: datetime | None) -> str:
    """e.g. '25 04 (Apr 25)'"""
    if not dt:
        return 'Unknown Date'
    return dt.strftime('%y %m (%b %y)')


def _date_prefix(dt: datetime | None) -> str:
    """e.g. '20250412'"""
    return dt.strftime('%Y%m%d') if dt else 'unknown'


def _build_thread_service(service):
    return build('gmail', 'v1', credentials=service._http.credentials)


def _html_to_text(html: str) -> str:
    if _HAS_BS4:
        soup = BeautifulSoup(html, 'html.parser')
        for tag in soup(['style', 'script', 'head']):
            tag.decompose()
        return soup.get_text(separator='\n').strip()
    text = re.sub(r'<[^>]+>', ' ', html)
    return re.sub(r' {2,}', ' ', text).strip()


# ─────────────────────────────────────────────────────────────────────────────
# GMAIL API HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _get_full_message(service, message_id: str) -> dict:
    return service.users().messages().get(
        userId='me', id=message_id, format='full'
    ).execute()


def _get_thread(service, thread_id: str) -> dict:
    return service.users().threads().get(
        userId='me', id=thread_id, format='full'
    ).execute()


def _headers_dict(message: dict) -> dict:
    headers = message.get('payload', {}).get('headers', [])
    return {h['name']: h['value'] for h in headers}


def _extract_body_and_attachments(payload: dict) -> tuple[str, list[dict]]:
    """Walk MIME tree. Returns (body_text, [{'filename', 'attachmentId', 'mime'}])"""
    body_text = ''
    body_html = ''
    attachments = []

    def _walk(part):
        nonlocal body_text, body_html
        mime = part.get('mimeType', '')
        body = part.get('body', {})
        if part.get('parts'):
            for sub in part['parts']:
                _walk(sub)
            return
        filename = part.get('filename', '')
        att_id   = body.get('attachmentId')
        if filename and att_id:
            attachments.append({'filename': filename, 'attachmentId': att_id, 'mime': mime})
            return
        data = body.get('data', '')
        if not data:
            return
        decoded = base64.urlsafe_b64decode(data + '==').decode('utf-8', errors='replace')
        if mime == 'text/plain' and not body_text:
            body_text = decoded
        elif mime == 'text/html' and not body_html:
            body_html = decoded

    _walk(payload)
    final_body = body_text if body_text else (_html_to_text(body_html) if body_html else '')
    return final_body, attachments


def _fetch_attachment_bytes(service, message_id: str, attachment_id: str) -> bytes:
    resp = service.users().messages().attachments().get(
        userId='me', messageId=message_id, id=attachment_id
    ).execute()
    return base64.urlsafe_b64decode(resp['data'] + '==')


# ─────────────────────────────────────────────────────────────────────────────
# ATTACHMENT CATEGORY
# ─────────────────────────────────────────────────────────────────────────────

_EXT_TO_CATEGORY = {
    '.xlsx': 'Excel',    '.xls': 'Excel',    '.csv': 'Excel',
    '.docx': 'Document', '.doc': 'Document', '.rtf': 'Document',
    '.txt':  'Document', '.md':  'Document',
    '.pdf':  'PDF',
    '.pptx': 'PPT',      '.ppt': 'PPT',      '.odp': 'PPT',
    '.jpg':  'Image',    '.jpeg':'Image',     '.png': 'Image',
    '.gif':  'Image',    '.bmp': 'Image',     '.tiff':'Image',
}

def _att_category(filename: str) -> str:
    ext = ('.' + filename.rsplit('.', 1)[-1].lower()) if '.' in filename else ''
    return _EXT_TO_CATEGORY.get(ext, 'Other')


# ─────────────────────────────────────────────────────────────────────────────
# PDF RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def _render_thread_pdf(messages: list[dict]) -> bytes:
    if not _HAS_REPORTLAB:
        raise RuntimeError("reportlab is not installed. Run: pip install reportlab")

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm,  bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', parent=styles['Heading1'],
                              textColor=colors.HexColor('#2E6B74'), fontSize=14)
    meta_s  = ParagraphStyle('M', parent=styles['Normal'], fontSize=9, leading=13)
    body_s  = ParagraphStyle('B', parent=styles['Normal'], fontSize=10, leading=14, spaceAfter=4)
    att_s   = ParagraphStyle('A', parent=styles['Normal'], fontSize=9, textColor=colors.grey)

    story = []
    for i, msg in enumerate(messages):
        hdrs    = _headers_dict(msg)
        subject = hdrs.get('Subject', '(no subject)')
        body, att_list = _extract_body_and_attachments(msg.get('payload', {}))

        if i == 0:
            story.append(Paragraph(subject, title_s))
        else:
            story.append(HRFlowable(width='100%', thickness=0.5, color=colors.lightgrey))

        story.append(Spacer(1, 6))
        for label, key in [('From', 'From'), ('To', 'To'), ('Date', 'Date')]:
            val = hdrs.get(key, '')
            safe_val = val.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
            story.append(Paragraph(f"<b>{label}:</b> {safe_val}", meta_s))
        story.append(Spacer(1, 8))

        for line in body.split('\n'):
            line = line.strip()
            if line:
                safe = line.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
                story.append(Paragraph(safe, body_s))

        if att_list:
            story.append(Spacer(1, 6))
            names = ', '.join(a['filename'] for a in att_list)
            story.append(Paragraph(f"Attachments: {names}", att_s))

        story.append(Spacer(1, 14))

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# METADATA EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _build_metadata_sheet(
    emails_data: list[dict],
    body_paths:  dict[str, str] | None,        # msg_id → zip-relative path
    att_paths:   dict[str, list[str]] | None,  # msg_id → [zip-relative paths]
) -> bytes:
    """
    Build the metadata Excel.

    Multi-attachment handling:
      Each attachment gets its own numbered row. The From/To/Date/Subject cells
      for that email are merged vertically across all its attachment rows, so the
      email metadata appears once on the left while each attachment link is on
      its own line on the right. This sidesteps Excel's one-hyperlink-per-cell
      limit cleanly.

    Hyperlinks are relative paths from the Excel file's location inside the ZIP,
    so they resolve correctly after extraction regardless of where the user saves
    the folder.
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Metadata"

    include_body = bool(body_paths)
    include_att  = bool(att_paths)

    base_headers  = ['From', 'To', 'Date', 'Subject']
    extra_headers = []
    if include_body: extra_headers.append('Email Body')
    if include_att:  extra_headers.append('Attachment')   # singular — one per row
    all_headers = base_headers + extra_headers
    N_BASE = len(base_headers)

    hdr_fill  = PatternFill('solid', fgColor='2E6B74')
    hdr_font  = Font(bold=True, color='FFFFFF')
    link_font = Font(color='1155CC', underline='single')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_align = Alignment(vertical='top', wrap_text=True)

    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center

    excel_row = 2
    for email in emails_data:
        msg_id   = email.get('id', '')
        att_list = (att_paths or {}).get(msg_id, [])
        # Number of rows this email occupies = max(1, number of attachments)
        n_rows   = max(1, len(att_list))

        # Write base metadata into the first row of this email's block
        for col, key in enumerate(['From', 'To', 'Date', 'Subject'], 1):
            cell = ws.cell(row=excel_row, column=col, value=email.get(key, ''))
            cell.alignment = top_align

        # Merge base columns vertically if this email spans multiple rows
        if n_rows > 1:
            for col in range(1, N_BASE + 1):
                ws.merge_cells(
                    start_row=excel_row, start_column=col,
                    end_row=excel_row + n_rows - 1, end_column=col,
                )

        # Email body link — merged across all rows for this email
        col_offset = N_BASE + 1
        if include_body:
            path = (body_paths or {}).get(msg_id)
            if path:
                cell = ws.cell(row=excel_row, column=col_offset, value='Open')
                cell.hyperlink = path
                cell.font      = link_font
                cell.alignment = top_align
            if n_rows > 1:
                ws.merge_cells(
                    start_row=excel_row, start_column=col_offset,
                    end_row=excel_row + n_rows - 1, end_column=col_offset,
                )
            col_offset += 1

        # Attachment links — one per row
        if include_att:
            for att_i, att_rel in enumerate(att_list):
                att_row  = excel_row + att_i
                filename = att_rel.rsplit('/', 1)[-1]
                label    = f"{att_i + 1}. {filename}"
                cell     = ws.cell(row=att_row, column=col_offset, value=label)
                cell.hyperlink = att_rel
                cell.font      = link_font
                cell.alignment = top_align
            if not att_list:
                pass  # leave blank

        excel_row += n_rows

    # Auto-width (based on header + sampled cell lengths)
    for col in ws.columns:
        vals    = [str(c.value or '') for c in col if c.value]
        max_len = max((len(v) for v in vals), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN EXPORT FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def build_export_zip_streaming(
    service,
    emails_data:         list[dict],
    selected_emails:     list[dict],
    include_metadata:    bool = True,
    include_attachments: bool = False,
    include_bodies:      bool = False,
    on_progress=None,    # callable(done, total, phase_label)
    stop_flag=None,      # callable() → bool
) -> bytes:
    """
    Build export ZIP with live progress reporting.

    on_progress(done, total, phase) is called after each email is processed.
    stop_flag() returning True aborts cleanly — returns whatever was built so far.
    """
    def _progress(done, total, phase):
        if on_progress:
            on_progress(done, total, phase)

    def _stopped():
        return stop_flag() if stop_flag else False

    svc         = _build_thread_service(service)
    export_date = datetime.now().strftime('%Y%m%d')
    root        = f"gmail_export_{export_date}"

    buf        = io.BytesIO()
    body_paths: dict[str, str]       = {}
    att_paths:  dict[str, list[str]] = {}
    total_sel  = len(selected_emails)

    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:

        # ── Attachments ───────────────────────────────────────────────────
        if include_attachments:
            for done_i, email_meta in enumerate(selected_emails):
                if _stopped(): break
                _progress(done_i, total_sel, "Fetching attachments")
                msg_id = email_meta['id']
                try:
                    full_msg = _get_full_message(svc, msg_id)
                except Exception:
                    continue

                hdrs     = _headers_dict(full_msg)
                email_dt = _parse_date(hdrs.get('Date', ''))
                month    = _month_folder(email_dt)
                dprefix  = _date_prefix(email_dt)
                _, attachments = _extract_body_and_attachments(full_msg.get('payload', {}))

                for att in attachments:
                    try:
                        att_bytes = _fetch_attachment_bytes(svc, msg_id, att['attachmentId'])
                    except Exception:
                        continue
                    category  = _att_category(att['filename'])
                    safe_name = _safe_filename(f"{dprefix}_{att['filename']}")
                    zip_path  = f"{root}/attachments/{month}/{category}/{safe_name}"
                    rel_path  = f"attachments/{month}/{category}/{safe_name}"
                    zf.writestr(zip_path, att_bytes)
                    att_paths.setdefault(msg_id, []).append(rel_path)

            _progress(total_sel, total_sel, "Attachments done")

        # ── Email bodies (PDF) ────────────────────────────────────────────
        if include_bodies and not _stopped():
            if not _HAS_REPORTLAB:
                raise RuntimeError("reportlab is not installed. Run: pip install reportlab")

            seen_threads: set[str] = set()
            thread_to_rel: dict[str, str] = {}

            for done_i, email_meta in enumerate(selected_emails):
                if _stopped(): break
                _progress(done_i, total_sel, "Rendering email content")
                msg_id    = email_meta['id']
                thread_id = email_meta.get('threadId', msg_id)

                if thread_id in seen_threads:
                    if thread_id in thread_to_rel:
                        body_paths[msg_id] = thread_to_rel[thread_id]
                    continue
                seen_threads.add(thread_id)

                messages = []
                try:
                    thread   = _get_thread(svc, thread_id)
                    messages = thread.get('messages', [])
                except Exception:
                    continue
                if not messages:
                    continue

                first_hdrs = _headers_dict(messages[0])
                subject    = first_hdrs.get('Subject', 'no_subject')
                thread_dt  = _parse_date(first_hdrs.get('Date', ''))
                month      = _month_folder(thread_dt)
                dprefix    = _date_prefix(thread_dt)
                filename   = _safe_filename(f"{dprefix}_{subject}") + '.pdf'
                zip_path   = f"{root}/email_content/{month}/{filename}"
                rel_path   = f"email_content/{month}/{filename}"

                try:
                    pdf_bytes = _render_thread_pdf(messages)
                except Exception as e:
                    pdf_bytes = f"Error: {e}".encode()
                    zip_path  = zip_path.replace('.pdf', '_ERROR.txt')
                    rel_path  = rel_path.replace('.pdf', '_ERROR.txt')

                zf.writestr(zip_path, pdf_bytes)
                thread_to_rel[thread_id] = rel_path

                for em in selected_emails:
                    if em.get('threadId', em['id']) == thread_id:
                        body_paths[em['id']] = rel_path

            _progress(total_sel, total_sel, "Email content done")

        # ── Metadata Excel ────────────────────────────────────────────────
        if include_metadata and not _stopped():
            _progress(0, 1, "Building metadata table")
            excel_bytes = _build_metadata_sheet(
                emails_data,
                body_paths if include_bodies      else None,
                att_paths  if include_attachments else None,
            )
            zf.writestr(f"{root}/metadata.xlsx", excel_bytes)
            _progress(1, 1, "Done")

    return buf.getvalue()


# Keep old name as alias so nothing else breaks
build_export_zip = build_export_zip_streaming
